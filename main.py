# ==================== IMPORTS ====================

from datetime import date, timedelta
from typing import List, Optional, Union
from fastapi import FastAPI, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from sqlalchemy import create_engine, Column, Integer, String, Date, DateTime, Float, Boolean, UniqueConstraint
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import io
import os
from datetime import datetime
import hashlib
import pandas as pd
from difflib import get_close_matches
import json

# ==================== BASE DE DATOS ====================

DATABASE_URL = os.environ.get("DATABASE_URL", "postgresql://localhost/registros")
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

engine = create_engine(DATABASE_URL, pool_pre_ping=True)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# ==================== UTILIDADES ====================

def hashear_pin(pin: str) -> str:
    return hashlib.sha256(pin.encode()).hexdigest()

# ==================== MODELOS ====================

class RegistroDB(Base):
    __tablename__ = "registros"
    id              = Column(Integer, primary_key=True, index=True)
    usuario         = Column(String, index=True)
    fecha           = Column(Date)
    tarea_principal = Column(String, index=True)
    subtarea        = Column(String)
    tiempo_minutos  = Column(Integer)
    proyecto        = Column(String, nullable=True)
    comentarios     = Column(String, nullable=True)
    created_at      = Column(DateTime, default=datetime.utcnow)
    updated_at      = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class UsuarioDB(Base):
    __tablename__ = "usuarios"
    id      = Column(Integer, primary_key=True, index=True)
    codigo  = Column(String, unique=True, index=True)
    nombre  = Column(String)
    pin     = Column(String)
    rol     = Column(String, default="operario")

class PlantillaDB(Base):
    __tablename__ = "plantillas"
    id         = Column(Integer, primary_key=True, index=True)
    nombre     = Column(String)
    usuario    = Column(String)
    created_at = Column(DateTime, default=datetime.utcnow)

class PlantillaItemDB(Base):
    __tablename__ = "plantilla_items"
    id              = Column(Integer, primary_key=True, index=True)
    plantilla_id    = Column(Integer)
    tarea_principal = Column(String)
    subtarea        = Column(String)
    tiempo_minutos  = Column(Integer)
    proyecto        = Column(String, nullable=True)
    comentarios     = Column(String, nullable=True)

class VolumenDB(Base):
    __tablename__ = "volumenes"
    id              = Column(Integer, primary_key=True, index=True)
    fecha           = Column(Date, index=True)
    tarea_principal = Column(String, index=True)
    unidades        = Column(Integer)
    horas_teoricas  = Column(Float)
    creado_por      = Column(String)
    comentarios     = Column(String, nullable=True)
    created_at      = Column(DateTime, default=datetime.utcnow)
    updated_at      = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class ObjetivoDB(Base):
    __tablename__ = "objetivos"
    id              = Column(Integer, primary_key=True, index=True)
    tarea_principal = Column(String, unique=True, index=True)
    uds_hora        = Column(Float)
    horas_jornada   = Column(Float, default=8.0)
    updated_at      = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class ConfiguracionDB(Base):
    __tablename__ = "configuracion"
    id             = Column(Integer, primary_key=True, index=True)
    clave          = Column(String, unique=True, index=True)
    valor          = Column(String)
    descripcion    = Column(String, nullable=True)
    updated_at     = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class TareaDB(Base):
    __tablename__ = "tareas"
    id              = Column(Integer, primary_key=True, index=True)
    nombre          = Column(String, unique=True, index=True)
    color           = Column(String, default="#94a3b8")
    activa          = Column(Boolean, default=True)
    created_at      = Column(DateTime, default=datetime.utcnow)

class SubtareaDB(Base):
    __tablename__ = "subtareas"
    id              = Column(Integer, primary_key=True, index=True)
    tarea_nombre    = Column(String, index=True)
    nombre          = Column(String)
    activa          = Column(Boolean, default=True)
    created_at      = Column(DateTime, default=datetime.utcnow)

class ProduccionDB(Base):
    __tablename__ = "produccion"
    id          = Column(Integer, primary_key=True, index=True)
    usuario     = Column(String, index=True)
    fecha       = Column(Date, index=True)
    subtarea    = Column(String, index=True)
    unidades    = Column(Float)
    created_at  = Column(DateTime, default=datetime.utcnow)
    updated_at  = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    __table_args__ = (
        UniqueConstraint("usuario", "fecha", "subtarea", name="uq_produccion_usuario_fecha_subtarea"),
    )

class ObjetivoSubtareaDB(Base):
    __tablename__ = "objetivos_subtarea"
    id              = Column(Integer, primary_key=True, index=True)
    subtarea        = Column(String, unique=True, index=True)
    uds_hora_target = Column(Float)
    updated_at      = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class HistorialImportacionDB(Base):
    __tablename__ = "historial_importaciones"
    id                 = Column(Integer, primary_key=True, index=True)
    nombre_fichero     = Column(String)
    usuario_carga      = Column(String)
    fecha_carga        = Column(DateTime, default=datetime.utcnow)
    filas_insertadas   = Column(Integer, default=0)
    filas_actualizadas = Column(Integer, default=0)
    filas_error        = Column(Integer, default=0)
    detalle_errores    = Column(String, nullable=True)

Base.metadata.create_all(bind=engine)
# ==================== INICIALIZACIÓN DE DATOS ====================

TAREAS_DEFAULT = [
    {"nombre": "Picking",   "color": "#3b82f6"},
    {"nombre": "Packing",   "color": "#8b5cf6"},
    {"nombre": "Inbound",   "color": "#f59e0b"},
    {"nombre": "Shipping",  "color": "#10b981"},
    {"nombre": "Ecommerce", "color": "#ef4444"},
]

SUBTAREAS_DEFAULT = {
    "Picking":   ["Picking Balda", "Picking Pallet", "Picking Percha", "Revisión de Picking",
                  "Picking Kardex", "Picking Recogepedidos", "Picking Obsoleto", "Inventario",
                  "Reposiciones", "Traspasos", "Recepción Kardex", "Formacion",
                  "Compactar", "Incidencias", "Varios IT", "Varios Trigo",
                  "Lanzar pedidos", "Reuniones", "Otros"],
    "Packing":   ["Mercado Piscina", "Mercado Contenedor", "Tienda Etiquetado", "Wholesale",
                  "Tienda RFID", "Tienda Empleado y Otros", "Materiales", "Runner",
                  "Formacion", "Incidencias", "Reuniones", "Otros", "Admin"],
    "Inbound":   ["Muelle", "Rec Pallet", "Rec Balda", "Rec Percha", "Rec Zapatos",
                  "Rec Trigo", "Devoluciones", "Compactar", "Recepción Kardex",
                  "Materiales", "Formacion", "Incidencias", "Reuniones", "Otros", "Admin"],
    "Shipping":  ["COURIER preparacion carga", "Courier Carga", "FW preparación carga",
                  "FW carga", "SERWELL carga", "Devoluciones", "Formacion",
                  "Incidencias", "Reuniones", "Inventario", "Otros"],
    "Ecommerce": ["Empaquetado", "Runner", "Store RQ Tienda", "Store RQ Tiendas",
                  "Calidad Devo", "Calidad OneStock", "Calidad Gestion", "Calidad Otro",
                  "Formacion", "Actividad", "Limpieza", "Otros", "Reuniones"],
}

SINONIMOS_COLUMNAS = {
    "usuario":  ["usuario", "operario", "trabajador", "empleado", "nombre", "worker",
                 "operator", "employee", "codigo", "code", "user"],
    "fecha":    ["fecha", "date", "dia", "day", "jornada", "fecha_trabajo"],
    "subtarea": ["subtarea", "subtask", "actividad", "activity", "tarea_detalle",
                 "operacion", "operation", "sub_tarea", "sub tarea"],
    "unidades": ["unidades", "units", "uds", "cantidad", "quantity", "ud",
                 "volumen", "piezas", "pieces", "bultos", "items"],
}

with SessionLocal() as session:

    if not session.query(UsuarioDB).filter(UsuarioDB.rol == "supervisor").first():
        session.add(UsuarioDB(
            codigo="ADMIN",
            nombre="Administrador",
            pin=hashear_pin("1234"),
            rol="supervisor"
        ))
        session.commit()

    for t in TAREAS_DEFAULT:
        if not session.query(ObjetivoDB).filter(ObjetivoDB.tarea_principal == t["nombre"]).first():
            session.add(ObjetivoDB(
                tarea_principal=t["nombre"],
                uds_hora=100.0,
                horas_jornada=8.0
            ))

    for t in TAREAS_DEFAULT:
        if not session.query(TareaDB).filter(TareaDB.nombre == t["nombre"]).first():
            session.add(TareaDB(nombre=t["nombre"], color=t["color"], activa=True))

    for tarea_nombre, subtareas in SUBTAREAS_DEFAULT.items():
        for sub in subtareas:
            if not session.query(SubtareaDB).filter(
                SubtareaDB.tarea_nombre == tarea_nombre,
                SubtareaDB.nombre == sub
            ).first():
                session.add(SubtareaDB(tarea_nombre=tarea_nombre, nombre=sub, activa=True))

    if not session.query(ConfiguracionDB).filter(ConfiguracionDB.clave == "minutos_jornada").first():
        session.add(ConfiguracionDB(
            clave="minutos_jornada",
            valor="480",
            descripcion="Minutos máximos por jornada laboral"
        ))

    todas_subtareas = session.query(SubtareaDB).all()
    for sub in todas_subtareas:
        if not session.query(ObjetivoSubtareaDB).filter(
            ObjetivoSubtareaDB.subtarea == sub.nombre
        ).first():
            session.add(ObjetivoSubtareaDB(
                subtarea=sub.nombre,
                uds_hora_target=0.0
            ))

    session.commit()

# ==================== DETECCIÓN AUTOMÁTICA DE COLUMNAS ====================

def detectar_columnas(columnas_fichero: list) -> dict:
    columnas_lower = {col: col.lower().strip() for col in columnas_fichero}
    resultado = {}
    usadas = set()

    for campo, sinonimos in SINONIMOS_COLUMNAS.items():
        mejor = None
        mejor_score = 0

        for col_original, col_lower in columnas_lower.items():
            if col_original in usadas:
                continue
            if col_lower in sinonimos:
                mejor = col_original
                mejor_score = 1.0
                break
            matches = get_close_matches(col_lower, sinonimos, n=1, cutoff=0.6)
            if matches:
                score = len(matches[0]) / max(len(col_lower), len(matches[0]))
                if score > mejor_score:
                    mejor = col_original
                    mejor_score = score

        if mejor:
            resultado[campo] = mejor
            usadas.add(mejor)

    resultado["no_mapeadas"] = [
        col for col in columnas_fichero if col not in usadas
    ]

    return resultado


def normalizar_subtarea(nombre: str, subtareas_validas: list) -> Optional[str]:
    if not nombre:
        return None
    nombre_lower = nombre.lower().strip()
    subtareas_lower = {s.lower().strip(): s for s in subtareas_validas}

    if nombre_lower in subtareas_lower:
        return subtareas_lower[nombre_lower]

    matches = get_close_matches(nombre_lower, subtareas_lower.keys(), n=1, cutoff=0.7)
    if matches:
        return subtareas_lower[matches[0]]

    return None
# ==================== APP Y MIDDLEWARES ====================

app = FastAPI(title="API Registro de Tareas", version="11.0.0")

ALLOWED_ORIGINS = [
    "https://proyecto-horas.onrender.com",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==================== SCHEMAS ====================

class RegistroTarea(BaseModel):
    usuario: str
    fecha: date
    tarea_principal: str
    subtarea: str
    tiempo_minutos: int
    proyecto: Optional[str] = None
    comentarios: Optional[str] = None

class RegistroTareaRespuesta(RegistroTarea):
    id: int
    created_at: Optional[datetime] = None
    updated_at: Optional[datetime] = None
    class Config:
        from_attributes = True

class UsuarioCreate(BaseModel):
    codigo: str
    nombre: str
    pin: str
    rol: str = "operario"

class UsuarioRespuesta(BaseModel):
    id: int
    codigo: str
    nombre: str
    rol: str
    class Config:
        from_attributes = True

class LoginRequest(BaseModel):
    codigo: str
    pin: str

class PlantillaItemCreate(BaseModel):
    tarea_principal: str
    subtarea: str
    tiempo_minutos: int
    proyecto: Optional[str] = None
    comentarios: Optional[str] = None

class PlantillaCreate(BaseModel):
    nombre: str
    usuario: str
    items: List[PlantillaItemCreate]

class VolumenCreate(BaseModel):
    fecha: date
    tarea_principal: str
    unidades: int
    horas_teoricas: float
    creado_por: str
    comentarios: Optional[str] = None

class VolumenRespuesta(VolumenCreate):
    id: int
    created_at: Optional[datetime] = None
    updated_at: Optional[datetime] = None
    class Config:
        from_attributes = True

class ObjetivoCreate(BaseModel):
    tarea_principal: str
    uds_hora: float
    horas_jornada: float = 8.0

class ObjetivoRespuesta(ObjetivoCreate):
    id: int
    updated_at: Optional[datetime] = None
    class Config:
        from_attributes = True

class TareaCreate(BaseModel):
    nombre: str
    color: str = "#94a3b8"
    activa: bool = True

class TareaRespuesta(TareaCreate):
    id: int
    class Config:
        from_attributes = True

class SubtareaCreate(BaseModel):
    tarea_nombre: str
    nombre: str
    activa: bool = True

class SubtareaRespuesta(SubtareaCreate):
    id: int
    class Config:
        from_attributes = True

class ConfiguracionUpdate(BaseModel):
    valor: str

class ConfiguracionRespuesta(BaseModel):
    clave: str
    valor: str
    descripcion: Optional[str] = None
    class Config:
        from_attributes = True

class ProduccionCreate(BaseModel):
    usuario: str
    fecha: date
    subtarea: str
    unidades: float

class ProduccionRespuesta(ProduccionCreate):
    id: int
    created_at: Optional[datetime] = None
    updated_at: Optional[datetime] = None
    class Config:
        from_attributes = True

class ObjetivoSubtareaCreate(BaseModel):
    subtarea: str
    uds_hora_target: float

class ObjetivoSubtareaRespuesta(ObjetivoSubtareaCreate):
    id: int
    updated_at: Optional[datetime] = None
    class Config:
        from_attributes = True

class ResultadoImportacion(BaseModel):
    nombre_fichero: str
    filas_procesadas: int
    filas_insertadas: int
    filas_actualizadas: int
    filas_error: int
    errores: List[dict] = []
    mapeo_columnas: dict = {}
    columnas_no_mapeadas: List[str] = []

class ProductividadRespuesta(BaseModel):
    usuario: str
    fecha: str
    subtarea: str
    unidades: float
    horas_reales: float
    uds_hora_real: Optional[float]
    uds_hora_target: Optional[float]
    pct_target: Optional[float]
    semaforo: str

class HistorialImportacionRespuesta(BaseModel):
    id: int
    nombre_fichero: str
    usuario_carga: str
    fecha_carga: datetime
    filas_insertadas: int
    filas_actualizadas: int
    filas_error: int
    detalle_errores: Optional[str] = None
    class Config:
        from_attributes = True

# ==================== DEPENDENCIAS ====================

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def get_minutos_jornada(db: Session) -> int:
    config = db.query(ConfiguracionDB).filter(
        ConfiguracionDB.clave == "minutos_jornada"
    ).first()
    return int(config.valor) if config else 480

def calcular_semaforo(pct: Optional[float]) -> str:
    if pct is None:
        return "sin_target"
    if pct >= 100:
        return "verde"
    if pct >= 80:
        return "ambar"
    return "rojo"

# ==================== RUTAS GENERALES ====================

@app.get("/")
def leer_index():
    ruta = os.path.join(os.path.dirname(__file__), "index.html")
    if os.path.exists(ruta):
        return FileResponse(ruta)
    return {"error": "No se encuentra index.html"}
# ==================== LOGIN ====================

@app.post("/login/")
def login(datos: LoginRequest, db: Session = Depends(get_db)):
    pin_hasheado = hashear_pin(datos.pin)
    usuario = db.query(UsuarioDB).filter(
        UsuarioDB.codigo == datos.codigo,
        UsuarioDB.pin == pin_hasheado
    ).first()
    if not usuario:
        raise HTTPException(status_code=401, detail="Código o PIN incorrectos")
    return {"id": usuario.id, "codigo": usuario.codigo, "nombre": usuario.nombre, "rol": usuario.rol}

# ==================== USUARIOS ====================

@app.post("/usuarios/", response_model=UsuarioRespuesta, status_code=201)
def crear_usuario(usuario: UsuarioCreate, db: Session = Depends(get_db)):
    if db.query(UsuarioDB).filter(UsuarioDB.codigo == usuario.codigo).first():
        raise HTTPException(status_code=400, detail="Ya existe un usuario con ese código")
    datos = usuario.dict()
    datos["pin"] = hashear_pin(datos["pin"])
    nuevo = UsuarioDB(**datos)
    db.add(nuevo)
    db.commit()
    db.refresh(nuevo)
    return nuevo

@app.get("/usuarios/", response_model=List[UsuarioRespuesta])
def obtener_usuarios(db: Session = Depends(get_db)):
    return db.query(UsuarioDB).all()

@app.delete("/usuarios/{id}", status_code=204)
def eliminar_usuario(id: int, db: Session = Depends(get_db)):
    usuario = db.query(UsuarioDB).filter(UsuarioDB.id == id).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    db.delete(usuario)
    db.commit()

# ==================== TAREAS ====================

@app.get("/tareas/", response_model=List[TareaRespuesta])
def obtener_tareas(solo_activas: bool = True, db: Session = Depends(get_db)):
    query = db.query(TareaDB)
    if solo_activas:
        query = query.filter(TareaDB.activa == True)
    return query.order_by(TareaDB.nombre).all()

@app.post("/tareas/", response_model=TareaRespuesta, status_code=201)
def crear_tarea(tarea: TareaCreate, db: Session = Depends(get_db)):
    if db.query(TareaDB).filter(TareaDB.nombre == tarea.nombre).first():
        raise HTTPException(status_code=400, detail="Ya existe una tarea con ese nombre")
    nueva = TareaDB(**tarea.dict())
    db.add(nueva)
    db.commit()
    db.refresh(nueva)
    return nueva

@app.put("/tareas/{id}", response_model=TareaRespuesta)
def editar_tarea(id: int, tarea: TareaCreate, db: Session = Depends(get_db)):
    db_tarea = db.query(TareaDB).filter(TareaDB.id == id).first()
    if not db_tarea:
        raise HTTPException(status_code=404, detail="Tarea no encontrada")
    for k, v in tarea.dict().items():
        setattr(db_tarea, k, v)
    db.commit()
    db.refresh(db_tarea)
    return db_tarea

@app.delete("/tareas/{id}", status_code=204)
def eliminar_tarea(id: int, db: Session = Depends(get_db)):
    db_tarea = db.query(TareaDB).filter(TareaDB.id == id).first()
    if not db_tarea:
        raise HTTPException(status_code=404, detail="Tarea no encontrada")
    db.delete(db_tarea)
    db.commit()

# ==================== SUBTAREAS ====================

@app.get("/subtareas/", response_model=List[SubtareaRespuesta])
def obtener_subtareas(tarea_nombre: Optional[str] = None, solo_activas: bool = True, db: Session = Depends(get_db)):
    query = db.query(SubtareaDB)
    if tarea_nombre:
        query = query.filter(SubtareaDB.tarea_nombre == tarea_nombre)
    if solo_activas:
        query = query.filter(SubtareaDB.activa == True)
    return query.order_by(SubtareaDB.tarea_nombre, SubtareaDB.nombre).all()

@app.post("/subtareas/", response_model=SubtareaRespuesta, status_code=201)
def crear_subtarea(subtarea: SubtareaCreate, db: Session = Depends(get_db)):
    if db.query(SubtareaDB).filter(
        SubtareaDB.tarea_nombre == subtarea.tarea_nombre,
        SubtareaDB.nombre == subtarea.nombre
    ).first():
        raise HTTPException(status_code=400, detail="Ya existe esa subtarea para esta tarea")
    nueva = SubtareaDB(**subtarea.dict())
    db.add(nueva)
    db.commit()
    db.refresh(nueva)
    if not db.query(ObjetivoSubtareaDB).filter(
        ObjetivoSubtareaDB.subtarea == subtarea.nombre
    ).first():
        db.add(ObjetivoSubtareaDB(subtarea=subtarea.nombre, uds_hora_target=0.0))
        db.commit()
    return nueva

@app.put("/subtareas/{id}", response_model=SubtareaRespuesta)
def editar_subtarea(id: int, subtarea: SubtareaCreate, db: Session = Depends(get_db)):
    db_sub = db.query(SubtareaDB).filter(SubtareaDB.id == id).first()
    if not db_sub:
        raise HTTPException(status_code=404, detail="Subtarea no encontrada")
    for k, v in subtarea.dict().items():
        setattr(db_sub, k, v)
    db.commit()
    db.refresh(db_sub)
    return db_sub

@app.delete("/subtareas/{id}", status_code=204)
def eliminar_subtarea(id: int, db: Session = Depends(get_db)):
    db_sub = db.query(SubtareaDB).filter(SubtareaDB.id == id).first()
    if not db_sub:
        raise HTTPException(status_code=404, detail="Subtarea no encontrada")
    db.delete(db_sub)
    db.commit()

# ==================== CONFIGURACIÓN ====================

@app.get("/configuracion/", response_model=List[ConfiguracionRespuesta])
def obtener_configuracion(db: Session = Depends(get_db)):
    return db.query(ConfiguracionDB).all()

@app.put("/configuracion/{clave}", response_model=ConfiguracionRespuesta)
def actualizar_configuracion(clave: str, datos: ConfiguracionUpdate, db: Session = Depends(get_db)):
    config = db.query(ConfiguracionDB).filter(ConfiguracionDB.clave == clave).first()
    if not config:
        raise HTTPException(status_code=404, detail="Configuración no encontrada")
    config.valor = datos.valor
    config.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(config)
    return config

# ==================== REGISTROS ====================

@app.post("/registros/", response_model=RegistroTareaRespuesta, status_code=201)
def crear_registro(registro: RegistroTarea, db: Session = Depends(get_db)):
    minutos_jornada = get_minutos_jornada(db)
    total = db.query(RegistroDB).filter(
        RegistroDB.usuario == registro.usuario,
        RegistroDB.fecha == registro.fecha
    ).with_entities(RegistroDB.tiempo_minutos).all()
    total_mins = sum(r.tiempo_minutos for r in total)
    if total_mins + registro.tiempo_minutos > minutos_jornada:
        raise HTTPException(
            status_code=400,
            detail=f"No se pueden superar {minutos_jornada} minutos. Minutos ya registrados: {total_mins}"
        )
    nuevo = RegistroDB(**registro.dict())
    db.add(nuevo)
    db.commit()
    db.refresh(nuevo)
    return nuevo

@app.get("/registros/", response_model=List[RegistroTareaRespuesta])
def obtener_registros(db: Session = Depends(get_db)):
    return db.query(RegistroDB).order_by(RegistroDB.fecha.desc()).all()

@app.get("/registros/rango/", response_model=List[RegistroTareaRespuesta])
def obtener_registros_rango(
    desde: date,
    hasta: date,
    usuario: Optional[str] = None,
    tarea: Optional[str] = None,
    db: Session = Depends(get_db)
):
    query = db.query(RegistroDB).filter(
        RegistroDB.fecha >= desde,
        RegistroDB.fecha <= hasta
    )
    if usuario:
        query = query.filter(RegistroDB.usuario == usuario)
    if tarea:
        query = query.filter(RegistroDB.tarea_principal == tarea)
    return query.order_by(RegistroDB.fecha.desc()).all()

@app.get("/registros/resumen-semanal/")
def resumen_semanal(
    usuario: Optional[str] = None,
    fecha_ref: Optional[date] = None,
    db: Session = Depends(get_db)
):
    if not fecha_ref:
        fecha_ref = date.today()
    inicio = fecha_ref - timedelta(days=fecha_ref.weekday())
    fin = inicio + timedelta(days=6)
    query = db.query(RegistroDB).filter(
        RegistroDB.fecha >= inicio,
        RegistroDB.fecha <= fin
    )
    if usuario:
        query = query.filter(RegistroDB.usuario == usuario)
    registros = query.all()
    resumen = {}
    for r in registros:
        dia = str(r.fecha)
        if dia not in resumen:
            resumen[dia] = {"total_minutos": 0, "tareas": {}}
        resumen[dia]["total_minutos"] += r.tiempo_minutos
        resumen[dia]["tareas"][r.tarea_principal] = (
            resumen[dia]["tareas"].get(r.tarea_principal, 0) + r.tiempo_minutos
        )
    return {
        "inicio": str(inicio),
        "fin": str(fin),
        "dias": resumen,
        "total_minutos": sum(r.tiempo_minutos for r in registros)
    }

@app.get("/registros/resumen-mensual/")
def resumen_mensual(
    usuario: Optional[str] = None,
    anyo: Optional[int] = None,
    mes: Optional[int] = None,
    db: Session = Depends(get_db)
):
    hoy = date.today()
    anyo = anyo or hoy.year
    mes = mes or hoy.month
    inicio = date(anyo, mes, 1)
    fin = (
        date(anyo + 1, 1, 1) - timedelta(days=1)
        if mes == 12
        else date(anyo, mes + 1, 1) - timedelta(days=1)
    )
    query = db.query(RegistroDB).filter(
        RegistroDB.fecha >= inicio,
        RegistroDB.fecha <= fin
    )
    if usuario:
        query = query.filter(RegistroDB.usuario == usuario)
    registros = query.all()
    por_usuario, por_tarea = {}, {}
    for r in registros:
        por_usuario[r.usuario] = por_usuario.get(r.usuario, 0) + r.tiempo_minutos
        por_tarea[r.tarea_principal] = por_tarea.get(r.tarea_principal, 0) + r.tiempo_minutos
    return {
        "anyo": anyo,
        "mes": mes,
        "total_minutos": sum(r.tiempo_minutos for r in registros),
        "total_registros": len(registros),
        "por_usuario": por_usuario,
        "por_tarea": por_tarea
    }

@app.get("/registros/{usuario}", response_model=List[RegistroTareaRespuesta])
def obtener_registros_por_usuario(
    usuario: str,
    desde: Optional[date] = None,
    hasta: Optional[date] = None,
    db: Session = Depends(get_db)
):
    query = db.query(RegistroDB).filter(RegistroDB.usuario.ilike(usuario))
    if desde:
        query = query.filter(RegistroDB.fecha >= desde)
    if hasta:
        query = query.filter(RegistroDB.fecha <= hasta)
    resultado = query.order_by(RegistroDB.fecha.desc()).all()
    if not resultado:
        raise HTTPException(status_code=404, detail="No se encontraron registros")
    return resultado

@app.put("/registros/{id}", response_model=RegistroTareaRespuesta)
def editar_registro(id: int, registro: RegistroTarea, db: Session = Depends(get_db)):
    minutos_jornada = get_minutos_jornada(db)
    db_reg = db.query(RegistroDB).filter(RegistroDB.id == id).first()
    if not db_reg:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    total = db.query(RegistroDB).filter(
        RegistroDB.usuario == registro.usuario,
        RegistroDB.fecha == registro.fecha,
        RegistroDB.id != id
    ).with_entities(RegistroDB.tiempo_minutos).all()
    total_mins = sum(r.tiempo_minutos for r in total)
    if total_mins + registro.tiempo_minutos > minutos_jornada:
        raise HTTPException(
            status_code=400,
            detail=f"No se pueden superar {minutos_jornada} minutos. Minutos ya registrados: {total_mins}"
        )
    for k, v in registro.dict().items():
        setattr(db_reg, k, v)
    db_reg.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(db_reg)
    return db_reg

@app.delete("/registros/{id}", status_code=204)
def eliminar_registro(id: int, db: Session = Depends(get_db)):
    db_reg = db.query(RegistroDB).filter(RegistroDB.id == id).first()
    if not db_reg:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    db.delete(db_reg)
    db.commit()

@app.post("/registros/{id}/duplicar", response_model=RegistroTareaRespuesta, status_code=201)
def duplicar_registro(id: int, nueva_fecha: Optional[date] = None, db: Session = Depends(get_db)):
    minutos_jornada = get_minutos_jornada(db)
    original = db.query(RegistroDB).filter(RegistroDB.id == id).first()
    if not original:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    fecha_destino = nueva_fecha or original.fecha
    total = db.query(RegistroDB).filter(
        RegistroDB.usuario == original.usuario,
        RegistroDB.fecha == fecha_destino
    ).with_entities(RegistroDB.tiempo_minutos).all()
    total_mins = sum(r.tiempo_minutos for r in total)
    if total_mins + original.tiempo_minutos > minutos_jornada:
        raise HTTPException(
            status_code=400,
            detail=f"No se pueden superar {minutos_jornada} minutos. Minutos ya registrados: {total_mins}"
        )
    nuevo = RegistroDB(
        usuario=original.usuario,
        fecha=fecha_destino,
        tarea_principal=original.tarea_principal,
        subtarea=original.subtarea,
        tiempo_minutos=original.tiempo_minutos,
        proyecto=original.proyecto,
        comentarios=original.comentarios
    )
    db.add(nuevo)
    db.commit()
    db.refresh(nuevo)
    return nuevo
# ==================== PLANTILLAS ====================

@app.post("/plantillas/", status_code=201)
def crear_plantilla(plantilla: PlantillaCreate, db: Session = Depends(get_db)):
    nueva = PlantillaDB(nombre=plantilla.nombre, usuario=plantilla.usuario)
    db.add(nueva)
    db.commit()
    db.refresh(nueva)
    for item in plantilla.items:
        db.add(PlantillaItemDB(
            plantilla_id=nueva.id,
            tarea_principal=item.tarea_principal,
            subtarea=item.subtarea,
            tiempo_minutos=item.tiempo_minutos,
            proyecto=item.proyecto,
            comentarios=item.comentarios
        ))
    db.commit()
    return {"id": nueva.id, "nombre": nueva.nombre, "mensaje": "Plantilla creada correctamente"}

@app.get("/plantillas/")
def obtener_plantillas(usuario: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(PlantillaDB)
    if usuario:
        query = query.filter(PlantillaDB.usuario == usuario)
    resultado = []
    for p in query.all():
        items = db.query(PlantillaItemDB).filter(PlantillaItemDB.plantilla_id == p.id).all()
        resultado.append({
            "id": p.id,
            "nombre": p.nombre,
            "usuario": p.usuario,
            "created_at": str(p.created_at),
            "items": [
                {
                    "id": i.id,
                    "tarea_principal": i.tarea_principal,
                    "subtarea": i.subtarea,
                    "tiempo_minutos": i.tiempo_minutos,
                    "proyecto": i.proyecto,
                    "comentarios": i.comentarios
                }
                for i in items
            ]
        })
    return resultado

@app.delete("/plantillas/{id}", status_code=204)
def eliminar_plantilla(id: int, db: Session = Depends(get_db)):
    plantilla = db.query(PlantillaDB).filter(PlantillaDB.id == id).first()
    if not plantilla:
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")
    db.query(PlantillaItemDB).filter(PlantillaItemDB.plantilla_id == id).delete()
    db.delete(plantilla)
    db.commit()

# ==================== OBJETIVOS POR TAREA ====================

@app.get("/objetivos/", response_model=List[ObjetivoRespuesta])
def obtener_objetivos(db: Session = Depends(get_db)):
    return db.query(ObjetivoDB).all()

@app.put("/objetivos/{tarea}", response_model=ObjetivoRespuesta)
def actualizar_objetivo(tarea: str, objetivo: ObjetivoCreate, db: Session = Depends(get_db)):
    db_obj = db.query(ObjetivoDB).filter(ObjetivoDB.tarea_principal == tarea).first()
    if not db_obj:
        db_obj = ObjetivoDB(
            tarea_principal=tarea,
            uds_hora=objetivo.uds_hora,
            horas_jornada=objetivo.horas_jornada
        )
        db.add(db_obj)
    else:
        db_obj.uds_hora = objetivo.uds_hora
        db_obj.horas_jornada = objetivo.horas_jornada
        db_obj.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(db_obj)
    return db_obj

# ==================== OBJETIVOS POR SUBTAREA (TARGETS) ====================

@app.get("/objetivos-subtarea/", response_model=List[ObjetivoSubtareaRespuesta])
def obtener_objetivos_subtarea(db: Session = Depends(get_db)):
    return db.query(ObjetivoSubtareaDB).order_by(ObjetivoSubtareaDB.subtarea).all()

@app.put("/objetivos-subtarea/{subtarea}", response_model=ObjetivoSubtareaRespuesta)
def actualizar_objetivo_subtarea(subtarea: str, objetivo: ObjetivoSubtareaCreate, db: Session = Depends(get_db)):
    db_obj = db.query(ObjetivoSubtareaDB).filter(
        ObjetivoSubtareaDB.subtarea == subtarea
    ).first()
    if not db_obj:
        db_obj = ObjetivoSubtareaDB(
            subtarea=subtarea,
            uds_hora_target=objetivo.uds_hora_target
        )
        db.add(db_obj)
    else:
        db_obj.uds_hora_target = objetivo.uds_hora_target
        db_obj.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(db_obj)
    return db_obj
@app.post("/objetivos-subtarea/bulk/")
def actualizar_objetivos_subtarea_bulk(
    request_data: Union[List[ObjetivoSubtareaCreate], dict],
    db: Session = Depends(get_db)
):
    actualizados = 0
    errores = []

    # Normalizar input a lista de dicts simples
    if isinstance(request_data, dict):
        elementos_dict = {}
        for k, v in request_data.items():
            partes = k.rsplit("__", 1)
            if len(partes) == 2 and partes[1].isdigit():
                campo, idx = partes[0], int(partes[1])
                if idx not in elementos_dict:
                    elementos_dict[idx] = {}
                elementos_dict[idx][campo] = v
        objetivos_raw = [
            {
                "subtarea": elementos_dict[i].get("subtarea"),
                "uds_hora_target": float(elementos_dict[i].get("uds_hora_target", 0.0))
            }
            for i in sorted(elementos_dict.keys())
            if elementos_dict[i].get("subtarea")
        ]
    else:
        objetivos_raw = [
            {"subtarea": o.subtarea, "uds_hora_target": o.uds_hora_target}
            for o in request_data
        ]

    try:
        for obj in objetivos_raw:
            # Usar SQL directo para evitar el bug de SQLAlchemy con bulk params
            db.execute(
                __import__("sqlalchemy").text(
                    "UPDATE objetivos_subtarea SET uds_hora_target = :uds, updated_at = :ts WHERE subtarea = :sub"
                ),
                {
                    "uds": obj["uds_hora_target"],
                    "ts": datetime.utcnow(),
                    "sub": obj["subtarea"]
                }
            )
            actualizados += 1

        db.commit()

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error en actualización masiva: {str(e)}")

    return {
        "actualizados": actualizados,
        "errores": errores
    }
# ==================== PRODUCCIÓN IMPORTADA ====================

@app.get("/produccion/", response_model=List[ProduccionRespuesta])
def obtener_produccion(
    desde: Optional[date] = None,
    hasta: Optional[date] = None,
    usuario: Optional[str] = None,
    subtarea: Optional[str] = None,
    db: Session = Depends(get_db)
):
    query = db.query(ProduccionDB)
    if desde:
        query = query.filter(ProduccionDB.fecha >= desde)
    if hasta:
        query = query.filter(ProduccionDB.fecha <= hasta)
    if usuario:
        query = query.filter(ProduccionDB.usuario == usuario)
    if subtarea:
        query = query.filter(ProduccionDB.subtarea == subtarea)
    return query.order_by(ProduccionDB.fecha.desc()).all()

@app.post("/produccion/importar/")
async def importar_produccion(
    file: UploadFile = File(...),
    usuario_carga: str = "ADMIN",
    db: Session = Depends(get_db)
):
    nombre = file.filename or ""
    extension = nombre.split(".")[-1].lower()
    if extension not in ["xlsx", "xls", "csv"]:
        raise HTTPException(
            status_code=400,
            detail="Formato no soportado. Use Excel (.xlsx, .xls) o CSV (.csv)"
        )

    contenido = await file.read()

    try:
        if extension == "csv":
            for sep in [",", ";", "\t", "|"]:
                try:
                    df = pd.read_csv(io.BytesIO(contenido), sep=sep)
                    if len(df.columns) > 1:
                        break
                except Exception:
                    continue
        else:
            df = pd.read_excel(io.BytesIO(contenido))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error leyendo el fichero: {str(e)}")

    if df.empty:
        raise HTTPException(status_code=400, detail="El fichero está vacío")

    df.columns = [str(c).strip() for c in df.columns]
    mapeo = detectar_columnas(list(df.columns))

    campos_requeridos = ["usuario", "fecha", "subtarea", "unidades"]
    campos_faltantes = [c for c in campos_requeridos if c not in mapeo]
    if campos_faltantes:
        raise HTTPException(
            status_code=400,
            detail=f"No se pudieron detectar las columnas: {campos_faltantes}. "
                   f"Columnas encontradas: {list(df.columns)}"
        )

    subtareas_validas = [
        s.nombre for s in db.query(SubtareaDB).filter(SubtareaDB.activa == True).all()
    ]

    usuarios_validos = {
        u.nombre.lower().strip(): u.codigo
        for u in db.query(UsuarioDB).all()
    }
    usuarios_validos.update({
        u.codigo.lower().strip(): u.codigo
        for u in db.query(UsuarioDB).all()
    })

    insertadas = 0
    actualizadas = 0
    errores = []

    for idx, row in df.iterrows():
        fila_num = idx + 2
        try:
            val_usuario  = str(row[mapeo["usuario"]]).strip()
            val_fecha    = row[mapeo["fecha"]]
            val_subtarea = str(row[mapeo["subtarea"]]).strip()
            val_unidades = row[mapeo["unidades"]]

            usuario_codigo = usuarios_validos.get(val_usuario.lower().strip())
            if not usuario_codigo:
                errores.append({
                    "fila": fila_num,
                    "error": f"Operario '{val_usuario}' no encontrado en el sistema"
                })
                continue

            try:
                if isinstance(val_fecha, str):
                    for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"]:
                        try:
                            val_fecha = datetime.strptime(val_fecha.strip(), fmt).date()
                            break
                        except ValueError:
                            continue
                elif hasattr(val_fecha, "date"):
                    val_fecha = val_fecha.date()
                elif isinstance(val_fecha, date):
                    pass
                else:
                    raise ValueError("Formato de fecha no reconocido")
            except Exception:
                errores.append({
                    "fila": fila_num,
                    "error": f"Fecha '{val_fecha}' no válida. Use formato YYYY-MM-DD o DD/MM/YYYY"
                })
                continue

            subtarea_normalizada = normalizar_subtarea(val_subtarea, subtareas_validas)
            if not subtarea_normalizada:
                errores.append({
                    "fila": fila_num,
                    "error": f"Subtarea '{val_subtarea}' no encontrada en el sistema"
                })
                continue

            try:
                unidades = float(val_unidades)
                if unidades < 0:
                    raise ValueError("Unidades negativas")
            except Exception:
                errores.append({
                    "fila": fila_num,
                    "error": f"Unidades '{val_unidades}' no válidas"
                })
                continue

            existente = db.query(ProduccionDB).filter(
                ProduccionDB.usuario  == usuario_codigo,
                ProduccionDB.fecha    == val_fecha,
                ProduccionDB.subtarea == subtarea_normalizada
            ).first()

            if existente:
                existente.unidades   = unidades
                existente.updated_at = datetime.utcnow()
                actualizadas += 1
            else:
                nuevo = ProduccionDB(
                    usuario  = usuario_codigo,
                    fecha    = val_fecha,
                    subtarea = subtarea_normalizada,
                    unidades = unidades
                )
                db.add(nuevo)
                insertadas += 1

        except Exception as e:
            errores.append({"fila": fila_num, "error": str(e)})

    db.commit()

    historial = HistorialImportacionDB(
        nombre_fichero     = nombre,
        usuario_carga      = usuario_carga,
        filas_insertadas   = insertadas,
        filas_actualizadas = actualizadas,
        filas_error        = len(errores),
        detalle_errores    = json.dumps(errores, ensure_ascii=False) if errores else None
    )
    db.add(historial)
    db.commit()

    return ResultadoImportacion(
        nombre_fichero       = nombre,
        filas_procesadas     = len(df),
        filas_insertadas     = insertadas,
        filas_actualizadas   = actualizadas,
        filas_error          = len(errores),
        errores              = errores,
        mapeo_columnas       = {k: v for k, v in mapeo.items() if k != "no_mapeadas"},
        columnas_no_mapeadas = mapeo.get("no_mapeadas", [])
    )

@app.get("/produccion/historial/", response_model=List[HistorialImportacionRespuesta])
def obtener_historial_importaciones(db: Session = Depends(get_db)):
    return db.query(HistorialImportacionDB).order_by(
        HistorialImportacionDB.fecha_carga.desc()
    ).all()
# ==================== PRODUCTIVIDAD ====================

@app.get("/productividad/")
def obtener_productividad(
    desde: Optional[date] = None,
    hasta: Optional[date] = None,
    usuario: Optional[str] = None,
    subtarea: Optional[str] = None,
    db: Session = Depends(get_db)
):
    if not desde:
        desde = date.today() - timedelta(days=30)
    if not hasta:
        hasta = date.today()

    query_prod = db.query(ProduccionDB).filter(
        ProduccionDB.fecha >= desde,
        ProduccionDB.fecha <= hasta
    )
    if usuario:
        query_prod = query_prod.filter(ProduccionDB.usuario == usuario)
    if subtarea:
        query_prod = query_prod.filter(ProduccionDB.subtarea == subtarea)

    producciones = query_prod.all()

    targets = {
        t.subtarea: t.uds_hora_target
        for t in db.query(ObjetivoSubtareaDB).all()
    }

    resultado = []
    for p in producciones:
        horas_reales = sum(
            r.tiempo_minutos for r in db.query(RegistroDB).filter(
                RegistroDB.usuario  == p.usuario,
                RegistroDB.fecha    == p.fecha,
                RegistroDB.subtarea == p.subtarea
            ).all()
        ) / 60

        uds_hora_real = round(p.unidades / horas_reales, 2) if horas_reales > 0 else None

        target = targets.get(p.subtarea)
        uds_hora_target = target if target and target > 0 else None

        pct_target = None
        if uds_hora_real is not None and uds_hora_target:
            pct_target = round((uds_hora_real / uds_hora_target) * 100, 1)

        resultado.append(ProductividadRespuesta(
            usuario         = p.usuario,
            fecha           = str(p.fecha),
            subtarea        = p.subtarea,
            unidades        = p.unidades,
            horas_reales    = round(horas_reales, 2),
            uds_hora_real   = uds_hora_real,
            uds_hora_target = uds_hora_target,
            pct_target      = pct_target,
            semaforo        = calcular_semaforo(pct_target)
        ))

    return sorted(resultado, key=lambda x: (x.fecha, x.usuario, x.subtarea), reverse=True)

@app.get("/productividad/resumen/")
def resumen_productividad(
    desde: Optional[date] = None,
    hasta: Optional[date] = None,
    db: Session = Depends(get_db)
):
    if not desde:
        desde = date.today() - timedelta(days=30)
    if not hasta:
        hasta = date.today()

    producciones = db.query(ProduccionDB).filter(
        ProduccionDB.fecha >= desde,
        ProduccionDB.fecha <= hasta
    ).all()

    targets = {
        t.subtarea: t.uds_hora_target
        for t in db.query(ObjetivoSubtareaDB).all()
    }

    por_subtarea = {}
    for p in producciones:
        horas_reales = sum(
            r.tiempo_minutos for r in db.query(RegistroDB).filter(
                RegistroDB.usuario  == p.usuario,
                RegistroDB.fecha    == p.fecha,
                RegistroDB.subtarea == p.subtarea
            ).all()
        ) / 60

        if p.subtarea not in por_subtarea:
            por_subtarea[p.subtarea] = {
                "unidades_total": 0,
                "horas_total": 0,
                "registros": 0
            }
        por_subtarea[p.subtarea]["unidades_total"] += p.unidades
        por_subtarea[p.subtarea]["horas_total"]    += horas_reales
        por_subtarea[p.subtarea]["registros"]      += 1

    resultado = []
    for sub, datos in por_subtarea.items():
        uds_hora_real = round(
            datos["unidades_total"] / datos["horas_total"], 2
        ) if datos["horas_total"] > 0 else None

        target = targets.get(sub)
        uds_hora_target = target if target and target > 0 else None

        pct_target = None
        if uds_hora_real and uds_hora_target:
            pct_target = round((uds_hora_real / uds_hora_target) * 100, 1)

        resultado.append({
            "subtarea":        sub,
            "unidades_total":  datos["unidades_total"],
            "horas_total":     round(datos["horas_total"], 2),
            "uds_hora_real":   uds_hora_real,
            "uds_hora_target": uds_hora_target,
            "pct_target":      pct_target,
            "semaforo":        calcular_semaforo(pct_target),
            "registros":       datos["registros"]
        })

    return sorted(resultado, key=lambda x: (x["pct_target"] or 0), reverse=True)

# ==================== VOLÚMENES ====================

@app.post("/volumenes/", response_model=VolumenRespuesta, status_code=201)
def crear_volumen(volumen: VolumenCreate, db: Session = Depends(get_db)):
    nuevo = VolumenDB(**volumen.dict())
    db.add(nuevo)
    db.commit()
    db.refresh(nuevo)
    return nuevo

@app.get("/volumenes/", response_model=List[VolumenRespuesta])
def obtener_volumenes(
    desde: Optional[date] = None,
    hasta: Optional[date] = None,
    tarea: Optional[str] = None,
    db: Session = Depends(get_db)
):
    query = db.query(VolumenDB)
    if desde:
        query = query.filter(VolumenDB.fecha >= desde)
    if hasta:
        query = query.filter(VolumenDB.fecha <= hasta)
    if tarea:
        query = query.filter(VolumenDB.tarea_principal == tarea)
    return query.order_by(VolumenDB.fecha.desc()).all()

@app.put("/volumenes/{id}", response_model=VolumenRespuesta)
def editar_volumen(id: int, volumen: VolumenCreate, db: Session = Depends(get_db)):
    db_vol = db.query(VolumenDB).filter(VolumenDB.id == id).first()
    if not db_vol:
        raise HTTPException(status_code=404, detail="Volumen no encontrado")
    for k, v in volumen.dict().items():
        setattr(db_vol, k, v)
    db_vol.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(db_vol)
    return db_vol

@app.delete("/volumenes/{id}", status_code=204)
def eliminar_volumen(id: int, db: Session = Depends(get_db)):
    db_vol = db.query(VolumenDB).filter(VolumenDB.id == id).first()
    if not db_vol:
        raise HTTPException(status_code=404, detail="Volumen no encontrado")
    db.delete(db_vol)
    db.commit()

@app.get("/volumenes/metricas/")
def obtener_metricas(
    desde: Optional[date] = None,
    hasta: Optional[date] = None,
    db: Session = Depends(get_db)
):
    if not desde:
        desde = date.today() - timedelta(days=30)
    if not hasta:
        hasta = date.today()
    volumenes = db.query(VolumenDB).filter(
        VolumenDB.fecha >= desde,
        VolumenDB.fecha <= hasta
    ).all()
    registros = db.query(RegistroDB).filter(
        RegistroDB.fecha >= desde,
        RegistroDB.fecha <= hasta
    ).all()
    resultado = []
    for v in volumenes:
        horas_reales = sum(
            r.tiempo_minutos for r in registros
            if str(r.fecha) == str(v.fecha) and r.tarea_principal == v.tarea_principal
        ) / 60
        eficiencia = round((v.horas_teoricas / horas_reales) * 100, 1) if horas_reales > 0 else None
        uds_hora_real = round(v.unidades / horas_reales, 1) if horas_reales > 0 else None
        uds_hora_teorica = round(v.unidades / v.horas_teoricas, 1) if v.horas_teoricas > 0 else None
        resultado.append({
            "id":               v.id,
            "fecha":            str(v.fecha),
            "tarea_principal":  v.tarea_principal,
            "unidades":         v.unidades,
            "horas_teoricas":   v.horas_teoricas,
            "horas_reales":     round(horas_reales, 2),
            "eficiencia_pct":   eficiencia,
            "uds_hora_real":    uds_hora_real,
            "uds_hora_teorica": uds_hora_teorica,
            "desviacion_horas": round(horas_reales - v.horas_teoricas, 2),
            "comentarios":      v.comentarios,
            "creado_por":       v.creado_por
        })
    return sorted(resultado, key=lambda x: x["fecha"], reverse=True)
# ==================== ESTADÍSTICAS ====================

@app.get("/estadisticas/productividad/")
def productividad(fecha_ref: Optional[date] = None, db: Session = Depends(get_db)):
    if not fecha_ref:
        fecha_ref = date.today()
    ayer = fecha_ref - timedelta(days=1)
    minutos_jornada = get_minutos_jornada(db)
    usuarios = db.query(UsuarioDB).filter(UsuarioDB.rol == "operario").all()
    resultado = []
    for u in usuarios:
        hoy_mins = db.query(RegistroDB).filter(
            RegistroDB.usuario == u.codigo,
            RegistroDB.fecha == fecha_ref
        ).with_entities(RegistroDB.tiempo_minutos).all()
        ayer_mins = db.query(RegistroDB).filter(
            RegistroDB.usuario == u.codigo,
            RegistroDB.fecha == ayer
        ).with_entities(RegistroDB.tiempo_minutos).all()
        total_hoy  = sum(r.tiempo_minutos for r in hoy_mins)
        total_ayer = sum(r.tiempo_minutos for r in ayer_mins)
        resultado.append({
            "usuario":            u.codigo,
            "nombre":             u.nombre,
            "minutos_hoy":        total_hoy,
            "minutos_ayer":       total_ayer,
            "variacion_minutos":  total_hoy - total_ayer,
            "porcentaje_jornada": round((total_hoy / minutos_jornada) * 100, 1)
        })
    return resultado

@app.get("/estadisticas/top-tareas/")
def top_tareas(
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    db: Session = Depends(get_db)
):
    if not fecha_desde:
        fecha_desde = date.today() - timedelta(days=30)
    if not fecha_hasta:
        fecha_hasta = date.today()
    registros = db.query(RegistroDB).filter(
        RegistroDB.fecha >= fecha_desde,
        RegistroDB.fecha <= fecha_hasta
    ).all()
    por_tarea = {}
    for r in registros:
        clave = f"{r.tarea_principal} - {r.subtarea}"
        if clave not in por_tarea:
            por_tarea[clave] = {"count": 0, "minutos": 0}
        por_tarea[clave]["count"]   += 1
        por_tarea[clave]["minutos"] += r.tiempo_minutos
    return [
        {
            "tarea":           k,
            "count":           v["count"],
            "minutos_totales": v["minutos"],
            "media_minutos":   round(v["minutos"] / v["count"], 1)
        }
        for k, v in sorted(por_tarea.items(), key=lambda x: x[1]["minutos"], reverse=True)[:10]
    ]

@app.get("/estadisticas/media-por-tarea/")
def media_por_tarea(usuario: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(RegistroDB)
    if usuario:
        query = query.filter(RegistroDB.usuario == usuario)
    por_tarea = {}
    for r in query.all():
        if r.tarea_principal not in por_tarea:
            por_tarea[r.tarea_principal] = {"count": 0, "minutos": 0}
        por_tarea[r.tarea_principal]["count"]   += 1
        por_tarea[r.tarea_principal]["minutos"] += r.tiempo_minutos
    return [
        {
            "tarea":         k,
            "count":         v["count"],
            "media_minutos": round(v["minutos"] / v["count"], 1),
            "total_minutos": v["minutos"]
        }
        for k, v in por_tarea.items()
    ]

# ==================== RENDIMIENTO ====================

@app.get("/rendimiento/ranking/")
def ranking_operarios(
    desde: Optional[date] = None,
    hasta: Optional[date] = None,
    db: Session = Depends(get_db)
):
    if not desde:
        desde = date.today() - timedelta(days=7)
    if not hasta:
        hasta = date.today()
    minutos_jornada = get_minutos_jornada(db)
    usuarios  = db.query(UsuarioDB).filter(UsuarioDB.rol == "operario").all()
    registros = db.query(RegistroDB).filter(
        RegistroDB.fecha >= desde,
        RegistroDB.fecha <= hasta
    ).all()
    dias_rango      = (hasta - desde).days + 1
    dias_laborables = sum(
        1 for i in range(dias_rango)
        if (desde + timedelta(days=i)).weekday() < 5
    )
    resultado = []
    for u in usuarios:
        regs_u     = [r for r in registros if r.usuario == u.codigo]
        total_mins = sum(r.tiempo_minutos for r in regs_u)
        dias_activo = len(set(str(r.fecha) for r in regs_u))
        max_mins    = dias_laborables * minutos_jornada
        pct_jornada = round((total_mins / max_mins) * 100, 1) if max_mins > 0 else 0
        consistencia = round((dias_activo / dias_laborables) * 100, 1) if dias_laborables > 0 else 0
        media_diaria = round(total_mins / dias_activo, 0) if dias_activo > 0 else 0
        resultado.append({
            "usuario":               u.codigo,
            "nombre":                u.nombre,
            "total_minutos":         total_mins,
            "total_horas":           round(total_mins / 60, 1),
            "dias_activo":           dias_activo,
            "dias_laborables":       dias_laborables,
            "pct_jornada_periodo":   pct_jornada,
            "consistencia_pct":      consistencia,
            "media_minutos_dia":     media_diaria
        })
    return sorted(resultado, key=lambda x: x["total_minutos"], reverse=True)

@app.get("/rendimiento/eficiencia-diaria/")
def eficiencia_diaria(
    desde: Optional[date] = None,
    hasta: Optional[date] = None,
    usuario: Optional[str] = None,
    db: Session = Depends(get_db)
):
    if not desde:
        desde = date.today() - timedelta(days=7)
    if not hasta:
        hasta = date.today()
    minutos_jornada = get_minutos_jornada(db)
    query = db.query(RegistroDB).filter(
        RegistroDB.fecha >= desde,
        RegistroDB.fecha <= hasta
    )
    if usuario:
        query = query.filter(RegistroDB.usuario == usuario)
    registros = query.all()
    por_dia = {}
    for r in registros:
        dia = str(r.fecha)
        if dia not in por_dia:
            por_dia[dia] = {"total_minutos": 0, "por_tarea": {}}
        por_dia[dia]["total_minutos"] += r.tiempo_minutos
        if r.tarea_principal not in por_dia[dia]["por_tarea"]:
            por_dia[dia]["por_tarea"][r.tarea_principal] = 0
        por_dia[dia]["por_tarea"][r.tarea_principal] += r.tiempo_minutos
    resultado = []
    for dia, datos in sorted(por_dia.items()):
        pct_jornada = round((datos["total_minutos"] / minutos_jornada) * 100, 1)
        resultado.append({
            "fecha":          dia,
            "total_minutos":  datos["total_minutos"],
            "total_horas":    round(datos["total_minutos"] / 60, 2),
            "pct_jornada":    pct_jornada,
            "por_tarea":      datos["por_tarea"]
        })
    return resultado

@app.get("/rendimiento/tendencia-semanal/")
def tendencia_semanal(
    semanas: Optional[int] = 4,
    usuario: Optional[str] = None,
    db: Session = Depends(get_db)
):
    hasta = date.today()
    desde = hasta - timedelta(weeks=semanas)
    query = db.query(RegistroDB).filter(
        RegistroDB.fecha >= desde,
        RegistroDB.fecha <= hasta
    )
    if usuario:
        query = query.filter(RegistroDB.usuario == usuario)
    registros    = query.all()
    semanas_data = {}
    for r in registros:
        iso   = r.fecha.isocalendar()
        clave = f"{iso[0]}-S{iso[1]:02d}"
        if clave not in semanas_data:
            semanas_data[clave] = {"minutos": 0, "dias": set(), "usuarios": set()}
        semanas_data[clave]["minutos"]  += r.tiempo_minutos
        semanas_data[clave]["dias"].add(str(r.fecha))
        semanas_data[clave]["usuarios"].add(r.usuario)
    return [
        {
            "semana":           k,
            "total_minutos":    v["minutos"],
            "total_horas":      round(v["minutos"] / 60, 1),
            "dias_activos":     len(v["dias"]),
            "usuarios_activos": len(v["usuarios"]),
            "media_horas_dia":  round(v["minutos"] / 60 / len(v["dias"]), 1) if v["dias"] else 0
        }
        for k, v in sorted(semanas_data.items())
    ]

@app.get("/rendimiento/alertas/")
def alertas_rendimiento(
    limite: Optional[int] = 5,
    solo_danger: Optional[bool] = False,
    db: Session = Depends(get_db)
):
    hoy            = date.today()
    hace_7         = hoy - timedelta(days=7)
    minutos_jornada = get_minutos_jornada(db)
    usuarios        = db.query(UsuarioDB).filter(UsuarioDB.rol == "operario").all()
    registros_hoy   = db.query(RegistroDB).filter(RegistroDB.fecha == hoy).all()
    registros_semana = db.query(RegistroDB).filter(
        RegistroDB.fecha >= hace_7,
        RegistroDB.fecha <= hoy
    ).all()

    grupos = {
        "sin_actividad":     {"nivel": "warning", "usuarios": [], "count": 0},
        "jornada_incompleta": {"nivel": "warning", "usuarios": [], "count": 0},
        "baja_consistencia":  {"nivel": "danger",  "usuarios": [], "count": 0},
        "eficiencia_baja":    {"nivel": "danger",  "subtareas": [], "count": 0},
        "desviacion":         {"nivel": "warning", "subtareas": [], "count": 0},
    }

    usuarios_con_actividad_hoy = set(r.usuario for r in registros_hoy)

    for u in usuarios:
        mins_hoy    = sum(r.tiempo_minutos for r in registros_hoy if r.usuario == u.codigo)
        mins_semana = sum(r.tiempo_minutos for r in registros_semana if r.usuario == u.codigo)
        dias_activo_semana = len(set(
            str(r.fecha) for r in registros_semana if r.usuario == u.codigo
        ))
        pct_hoy    = round((mins_hoy / minutos_jornada) * 100, 1)

        if u.codigo not in usuarios_con_actividad_hoy:
            grupos["sin_actividad"]["usuarios"].append(u.nombre)
            grupos["sin_actividad"]["count"] += 1

        if mins_hoy > 0 and pct_hoy < 70:
            grupos["jornada_incompleta"]["usuarios"].append(
                f"{u.nombre} ({pct_hoy}%)"
            )
            grupos["jornada_incompleta"]["count"] += 1

        if 0 < dias_activo_semana < 3:
            grupos["baja_consistencia"]["usuarios"].append(
                f"{u.nombre} ({dias_activo_semana} días)"
            )
            grupos["baja_consistencia"]["count"] += 1

    volumenes = db.query(VolumenDB).filter(
        VolumenDB.fecha >= hace_7,
        VolumenDB.fecha <= hoy
    ).all()
    for v in volumenes:
        horas_reales = sum(
            r.tiempo_minutos for r in registros_semana
            if str(r.fecha) == str(v.fecha) and r.tarea_principal == v.tarea_principal
        ) / 60
        if horas_reales > 0:
            eficiencia = round((v.horas_teoricas / horas_reales) * 100, 1)
            if eficiencia < 70:
                grupos["eficiencia_baja"]["subtareas"].append(
                    f"{v.tarea_principal} ({v.fecha}: {eficiencia}%)"
                )
                grupos["eficiencia_baja"]["count"] += 1
            elif abs(horas_reales - v.horas_teoricas) / v.horas_teoricas > 0.20:
                grupos["desviacion"]["subtareas"].append(
                    f"{v.tarea_principal} ({v.fecha})"
                )
                grupos["desviacion"]["count"] += 1

    MENSAJES = {
        "sin_actividad":      "operario(s) sin actividad hoy",
        "jornada_incompleta": "operario(s) con jornada incompleta",
        "baja_consistencia":  "operario(s) con baja consistencia esta semana",
        "eficiencia_baja":    "subtarea(s) con eficiencia baja",
        "desviacion":         "subtarea(s) con desviación >20%",
    }

    alertas_agrupadas = []
    for tipo, datos in grupos.items():
        if datos["count"] == 0:
            continue
        if solo_danger and datos["nivel"] != "danger":
            continue
        detalle = datos.get("usuarios") or datos.get("subtareas") or []
        alertas_agrupadas.append({
            "tipo":    tipo,
            "nivel":   datos["nivel"],
            "count":   datos["count"],
            "mensaje": f"{datos['count']} {MENSAJES[tipo]}",
            "detalle": detalle
        })

    alertas_agrupadas.sort(key=lambda x: 0 if x["nivel"] == "danger" else 1)
    alertas_limitadas = alertas_agrupadas[:limite]

    return {
        "total_grupos":  len(alertas_agrupadas),
        "total_alertas": sum(g["count"] for g in alertas_agrupadas),
        "mostrando":     len(alertas_limitadas),
        "alertas":       alertas_limitadas
    }

@app.get("/rendimiento/objetivo-vs-real/")
def objetivo_vs_real(
    desde: Optional[date] = None,
    hasta: Optional[date] = None,
    db: Session = Depends(get_db)
):
    if not desde:
        desde = date.today() - timedelta(days=30)
    if not hasta:
        hasta = date.today()
    registros = db.query(RegistroDB).filter(
        RegistroDB.fecha >= desde,
        RegistroDB.fecha <= hasta
    ).all()
    objetivos = {o.tarea_principal: o for o in db.query(ObjetivoDB).all()}
    por_tarea = {}
    for r in registros:
        if r.tarea_principal not in por_tarea:
            por_tarea[r.tarea_principal] = {"minutos": 0, "count": 0}
        por_tarea[r.tarea_principal]["minutos"] += r.tiempo_minutos
        por_tarea[r.tarea_principal]["count"]   += 1
    resultado = []
    for tarea, datos in por_tarea.items():
        obj          = objetivos.get(tarea)
        horas_reales = round(datos["minutos"] / 60, 2)
        horas_objetivo = round(obj.horas_jornada, 2) if obj else None
        pct = round((horas_reales / horas_objetivo) * 100, 1) if horas_objetivo else None
        resultado.append({
            "tarea":             tarea,
            "horas_reales":      horas_reales,
            "horas_objetivo":    horas_objetivo,
            "uds_hora_objetivo": obj.uds_hora if obj else None,
            "pct_objetivo":      pct,
            "registros":         datos["count"]
        })
    return sorted(resultado, key=lambda x: x["horas_reales"], reverse=True)

# ==================== EXCEL ====================

@app.get("/exportar-excel/")
def exportar_excel(
    desde: Optional[date] = None,
    hasta: Optional[date] = None,
    usuario: Optional[str] = None,
    db: Session = Depends(get_db)
):
    query = db.query(RegistroDB)
    if desde:
        query = query.filter(RegistroDB.fecha >= desde)
    if hasta:
        query = query.filter(RegistroDB.fecha <= hasta)
    if usuario:
        query = query.filter(RegistroDB.usuario == usuario)
    registros = query.order_by(RegistroDB.fecha.desc()).all()

    wb          = openpyxl.Workbook()
    HEADER_FILL  = PatternFill("solid", fgColor="1e293b")
    HEADER_FONT  = Font(bold=True, color="FFFFFF")
    HEADER_ALIGN = Alignment(horizontal="center")

    def estilo_cabecera(ws):
        for cell in ws[1]:
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = HEADER_ALIGN

    def autoajustar(ws):
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

    # Hoja 0: Resumen Ejecutivo
    ws0 = wb.active
    ws0.title = "Resumen Ejecutivo"
    ws0.append(["RESUMEN EJECUTIVO", ""])
    ws0["A1"].font = Font(bold=True, size=14, color="1e293b")
    ws0.append(["Generado el", str(date.today())])
    ws0.append(["Periodo", f"{desde or 'Todo'} → {hasta or 'Todo'}"])
    ws0.append(["Usuario", usuario or "Todos"])
    ws0.append([])
    ws0.append(["MÉTRICAS GLOBALES", ""])
    ws0["A6"].font = Font(bold=True, color="FFFFFF")
    ws0["A6"].fill = PatternFill("solid", fgColor="3b82f6")
    total_mins     = sum(r.tiempo_minutos for r in registros)
    usuarios_unicos = len(set(r.usuario for r in registros))
    dias_unicos    = len(set(str(r.fecha) for r in registros))
    ws0.append(["Total registros",    len(registros)])
    ws0.append(["Total horas",        round(total_mins / 60, 2)])
    ws0.append(["Usuarios activos",   usuarios_unicos])
    ws0.append(["Días con actividad", dias_unicos])
    ws0.append(["Media horas/día",    round(total_mins / 60 / dias_unicos, 2) if dias_unicos else 0])
    autoajustar(ws0)

    # Hoja 1: Registros Detallados
    ws1 = wb.create_sheet("Registros Detallados")
    ws1.append(["ID", "Usuario", "Fecha", "Semana", "Mes",
                "Tarea Principal", "Subtarea", "Tiempo (min)", "Horas", "Proyecto", "Comentarios"])
    estilo_cabecera(ws1)
    for r in registros:
        semana = r.fecha.isocalendar()[1]
        mes    = r.fecha.strftime("%B %Y")
        ws1.append([
            r.id, r.usuario, str(r.fecha), f"Semana {semana}", mes,
            r.tarea_principal, r.subtarea, r.tiempo_minutos,
            round(r.tiempo_minutos / 60, 2),
            r.proyecto or "", r.comentarios or ""
        ])
    ws1.auto_filter.ref = ws1.dimensions
    ws1.freeze_panes    = "A2"
    autoajustar(ws1)

    # Hoja 2: Resumen por Usuario
    ws2 = wb.create_sheet("Resumen por Usuario")
    ws2.append(["Usuario", "Total Registros", "Total Minutos",
                "Total Horas", "Días Activos", "Media Min/Día", "% Jornada Media"])
    estilo_cabecera(ws2)
    por_usuario     = {}
    minutos_jornada = get_minutos_jornada(db)
    for r in registros:
        if r.usuario not in por_usuario:
            por_usuario[r.usuario] = {"registros": 0, "minutos": 0, "dias": set()}
        por_usuario[r.usuario]["registros"] += 1
        por_usuario[r.usuario]["minutos"]   += r.tiempo_minutos
        por_usuario[r.usuario]["dias"].add(str(r.fecha))
    for u, d in por_usuario.items():
        dias      = len(d["dias"]) or 1
        media_dia = round(d["minutos"] / dias, 1)
        pct       = round((media_dia / minutos_jornada) * 100, 1)
        ws2.append([u, d["registros"], d["minutos"],
                    round(d["minutos"] / 60, 2), len(d["dias"]), media_dia, pct])
    autoajustar(ws2)

    # Hoja 3: Resumen por Tarea
    ws3 = wb.create_sheet("Resumen por Tarea")
    ws3.append(["Tarea Principal", "Subtarea", "Total Registros",
                "Total Minutos", "Total Horas", "Media Min/Registro"])
    estilo_cabecera(ws3)
    por_tarea = {}
    for r in registros:
        clave = (r.tarea_principal, r.subtarea)
        if clave not in por_tarea:
            por_tarea[clave] = {"registros": 0, "minutos": 0}
        por_tarea[clave]["registros"] += 1
        por_tarea[clave]["minutos"]   += r.tiempo_minutos
    for (tp, st), d in sorted(por_tarea.items()):
        media = round(d["minutos"] / d["registros"], 1) if d["registros"] > 0 else 0
        ws3.append([tp, st, d["registros"], d["minutos"], round(d["minutos"] / 60, 2), media])
    autoajustar(ws3)

    # Hoja 4: Productividad por Subtarea
    ws4 = wb.create_sheet("Productividad Subtarea")
    ws4.append(["Usuario", "Fecha", "Subtarea", "Unidades",
                "Horas Reales", "Ud/h Real", "Ud/h Target", "% Target", "Semáforo"])
    estilo_cabecera(ws4)
    producciones = db.query(ProduccionDB).order_by(ProduccionDB.fecha.desc()).all()
    targets_sub  = {t.subtarea: t.uds_hora_target for t in db.query(ObjetivoSubtareaDB).all()}
    SEMAFORO_TEXTO = {"verde": "✅ OK", "ambar": "⚠️ Mejorable", "rojo": "🔴 Bajo", "sin_target": "—"}
    for p in producciones:
        horas_reales = sum(
            r.tiempo_minutos for r in db.query(RegistroDB).filter(
                RegistroDB.usuario  == p.usuario,
                RegistroDB.fecha    == p.fecha,
                RegistroDB.subtarea == p.subtarea
            ).all()
        ) / 60
        uds_hora_real   = round(p.unidades / horas_reales, 2) if horas_reales > 0 else None
        target          = targets_sub.get(p.subtarea)
        uds_hora_target = target if target and target > 0 else None
        pct_target      = round((uds_hora_real / uds_hora_target) * 100, 1) if uds_hora_real and uds_hora_target else None
        semaforo        = calcular_semaforo(pct_target)
        ws4.append([
            p.usuario, str(p.fecha), p.subtarea, p.unidades,
            round(horas_reales, 2),
            uds_hora_real or 0,
            uds_hora_target or 0,
            pct_target or 0,
            SEMAFORO_TEXTO.get(semaforo, "—")
        ])
    autoajustar(ws4)

    # Hoja 5: Historial Importaciones
    ws5 = wb.create_sheet("Historial Importaciones")
    ws5.append(["ID", "Fichero", "Usuario Carga", "Fecha Carga",
                "Insertadas", "Actualizadas", "Errores"])
    estilo_cabecera(ws5)
    historial = db.query(HistorialImportacionDB).order_by(
        HistorialImportacionDB.fecha_carga.desc()
    ).all()
    for h in historial:
        ws5.append([
            h.id, h.nombre_fichero, h.usuario_carga,
            str(h.fecha_carga), h.filas_insertadas,
            h.filas_actualizadas, h.filas_error
        ])
    autoajustar(ws5)

    # Hoja 6: Ranking Rendimiento
    ws6 = wb.create_sheet("Ranking Rendimiento")
    ws6.append(["Posición", "Usuario", "Total Horas", "Días Activos",
                "% Jornada Periodo", "Consistencia %", "Media Min/Día"])
    estilo_cabecera(ws6)
    ranking = sorted(por_usuario.items(), key=lambda x: x[1]["minutos"], reverse=True)
    for pos, (u, d) in enumerate(ranking, 1):
        dias     = len(d["dias"]) or 1
        media    = round(d["minutos"] / dias, 1)
        pct_periodo = round((d["minutos"] / (dias * minutos_jornada)) * 100, 1)
        ws6.append([
            pos, u, round(d["minutos"] / 60, 1), len(d["dias"]),
            pct_periodo, round((dias / max(dias_unicos, 1)) * 100, 1), media
        ])
    autoajustar(ws6)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=registros.xlsx"}
    )

# ==================== ARRANQUE ====================

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run("main:app", host="0.0.0.0", port=port)
